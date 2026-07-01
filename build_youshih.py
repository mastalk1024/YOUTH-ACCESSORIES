import json, zipfile, datetime, re, os, sys
from pathlib import Path

# ─── Resolve folders relative to this script ──────────────────────────────────
SCRIPT_DIR   = Path(sys.argv[1])
YOUSHIH_DIR  = SCRIPT_DIR / '優仕'
WEIKU_DIR    = SCRIPT_DIR / '暐固'
OUTPUT_HTML  = SCRIPT_DIR / '配件產品目錄.html'

for d in (YOUSHIH_DIR, WEIKU_DIR):
    if not d.exists():
        print(f"⚠️  找不到資料夾：{d}（略過）")

# ─── Package check ────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("安裝 openpyxl..."); os.system(f"{sys.executable} -m pip install openpyxl -q"); import openpyxl
try:
    from lxml import etree
except ImportError:
    print("安裝 lxml..."); os.system(f"{sys.executable} -m pip install lxml -q"); from lxml import etree

# ─── Common helpers ───────────────────────────────────────────────────────────
def excel_date_1904(serial):
    try:
        base = datetime.date(1904, 1, 1)
        return (base + datetime.timedelta(days=int(float(serial)))).strftime('%Y-%m-%d')
    except:
        return ''

def excel_date_1900(serial):
    try:
        base = datetime.date(1899, 12, 30)
        return (base + datetime.timedelta(days=int(float(serial)))).strftime('%Y-%m-%d')
    except:
        return ''

def safe_int(v):
    if not v: return ''
    try: return str(int(float(str(v).replace(',',''))))
    except: return ''

def safe_str(v):
    if v is None: return ''
    s = str(v).strip()
    return '' if s in ('None','#N/A','N/A') else s

def get_val(cells, keys):
    for k in keys:
        v = cells.get(k)
        s = safe_str(v)
        if s: return v
    return None

def parse_shared_strings(z):
    shared = []
    if 'xl/sharedStrings.xml' in z.namelist():
        root = etree.fromstring(z.read('xl/sharedStrings.xml'))
        ns = {'s':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        for si in root.findall('s:si', ns):
            shared.append(''.join(si.itertext()))
    return shared

def parse_sheet_xml(z, sheet_file, shared):
    sp = f'xl/{sheet_file}'
    if sp not in z.namelist(): return []
    ns = {'s':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    ws_root = etree.fromstring(z.read(sp))
    rows = []
    for row_el in ws_root.findall('.//s:row', ns):
        cells = {}
        for c in row_el.findall('s:c', ns):
            ref = c.get('r','')
            col = ''.join(ch for ch in ref if ch.isalpha())
            t = c.get('t','')
            v_el = c.find('s:v', ns)
            val = None
            if v_el is not None and v_el.text:
                if t == 's':
                    idx = int(v_el.text)
                    val = shared[idx] if idx < len(shared) else ''
                elif t in ('str','inlineStr'): val = v_el.text
                else:
                    try:
                        f = float(v_el.text); val = int(f) if f == int(f) else f
                    except: val = v_el.text
            if val is not None: cells[col] = val
        if cells: rows.append(cells)
    return rows

def get_sheet_map(z):
    """Returns {sheet_name: (file_target, state)}"""
    wb_root = etree.fromstring(z.read('xl/workbook.xml'))
    rels_root = etree.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    ns_s = {'s':'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
            'r':'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
    ns_r = {'r':'http://schemas.openxmlformats.org/package/2006/relationships'}
    rid_file = {r.get('Id'): r.get('Target') for r in rels_root.findall('r:Relationship', ns_r)}
    result = {}
    for sh in wb_root.findall('.//s:sheet', ns_s):
        name = sh.get('name')
        state = sh.get('state','')
        rid = sh.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        result[name] = (rid_file.get(rid,''), state)
    return result

# ═══════════════════════════════════════════════════════════════════════════════
# 優仕 parser
# ═══════════════════════════════════════════════════════════════════════════════
YOUSHIH_SKIP = {
    '綜合表格','國條流水號','配件展示','工作表2','工作表1',
    '給優邑專用','CITIESOCIAL','商品券','神腦','優銳配件',
    '優仕包膜','虛擬服務類','贈品類','媒體儲存',
}
YS_HEADER_KW = {'Barcode','品名','優仕料號','料號','Y料號','僅國條','barcode'}

def ys_is_header(d):
    return bool(YS_HEADER_KW & set(safe_str(v) for v in d.values()))

def ys_detect_cols(hd):
    m = {'料號':[],'barcode':[],'品名':[],'進價':[],'售價':[],'廠商':[],'日期':[]}
    for col, val in hd.items():
        v = safe_str(val)
        if '料號' in v or v == '僅國條': m['料號'].append(col)
        elif 'Barcode' in v or 'barcode' in v or ('國條' in v and '僅' not in v): m['barcode'].append(col)
        elif '品名' in v: m['品名'].append(col)
        elif '進價' in v: m['進價'].append(col)
        elif '售價' in v: m['售價'].append(col)
        elif '廠商' in v: m['廠商'].append(col)
        elif '建檔' in v and '更新' not in v: m['日期'].append(col)
    return m

def ys_make_row(cells, cm, date1904=False):
    品名 = get_val(cells, cm['品名'])
    if not 品名: return None
    日期 = get_val(cells, cm['日期'])
    if isinstance(日期, datetime.datetime): 日期 = 日期.strftime('%Y-%m-%d')
    elif isinstance(日期, (int,float)) and date1904: 日期 = excel_date_1904(日期)
    elif isinstance(日期, (int,float)): 日期 = excel_date_1900(日期)
    elif 日期: 日期 = safe_str(日期)
    料號 = get_val(cells, cm['料號'])
    bc   = get_val(cells, cm['barcode'])
    return {
        '料號':    safe_str(料號),
        'barcode': safe_str(bc).replace('.0',''),
        '品名':    safe_str(品名),
        '進價':    safe_int(get_val(cells, cm['進價'])),
        '售價':    safe_int(get_val(cells, cm['售價'])),
        '廠商':    safe_str(get_val(cells, cm['廠商'])),
        '建檔日期': 日期 or '',
    }

def parse_youshih_openpyxl(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except:
        return None
    result = {}
    for sname in wb.sheetnames:
        if sname in YOUSHIH_SKIP: continue
        ws = wb[sname]
        if ws.sheet_state == 'hidden':
            print(f"    ⚠️  分頁「{sname}」是隱藏分頁，已略過（若這是正式分類請取消隱藏）")
            continue
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True, max_col=20))
        cm, dstart = None, 0
        for i, row in enumerate(rows[:5]):
            hd = {chr(65+j): v for j,v in enumerate(row) if v is not None}
            if ys_is_header(hd): cm = ys_detect_cols(hd); dstart = i+1; break
        if not cm:
            if len(rows) > 0:
                print(f"    ⚠️  分頁「{sname}」有 {len(rows)} 列資料但找不到表頭，已略過（請確認欄位標題是否正常）")
            continue
        prods = []
        for row in rows[dstart:]:
            cells = {}
            for j, v in enumerate(row):
                if isinstance(v, datetime.datetime): v = v.strftime('%Y-%m-%d')
                cells[chr(65+j)] = v
            r = ys_make_row(cells, cm, date1904=False)
            if r: prods.append(r)
        if prods: result[sname] = prods
    return result

def parse_youshih_xml(path):
    result = {}
    with zipfile.ZipFile(path, 'r') as z:
        shared = parse_shared_strings(z)
        sheet_map = get_sheet_map(z)
        for name, (sf, state) in sheet_map.items():
            if name in YOUSHIH_SKIP: continue
            if state == 'hidden':
                print(f"    ⚠️  分頁「{name}」是隱藏分頁，已略過（若這是正式分類請取消隱藏）")
                continue
            rows_raw = parse_sheet_xml(z, sf, shared)
            cm, dstart = None, 0
            for i, cells in enumerate(rows_raw[:5]):
                if ys_is_header(cells): cm = ys_detect_cols(cells); dstart = i+1; break
            if not cm:
                if len(rows_raw) > 0:
                    print(f"    ⚠️  分頁「{name}」有 {len(rows_raw)} 列資料但找不到表頭，已略過（請確認欄位標題是否正常）")
                continue
            prods = []
            for cells in rows_raw[dstart:]:
                r = ys_make_row(cells, cm, date1904=True)
                if r: prods.append(r)
            if prods: result[name] = prods
    return result

# ═══════════════════════════════════════════════════════════════════════════════
# 暐固 parser
# ═══════════════════════════════════════════════════════════════════════════════
WEIKU_SKIP = {'未進貨','席德曼','展示機'}
WEIKU_RENAME = {'工作表1': '保護類'}  # sheet 顯示名稱對照
WK_HEADER_KW = {'料號','國際條碼','品名','供應商','品牌','T1成本','T1價格'}

def wk_is_header(d):
    vals = set(safe_str(v).split('\n')[0] for v in d.values())
    return bool(WK_HEADER_KW & vals) or any('品名' in safe_str(v) for v in d.values())

def wk_detect_cols(hd):
    m = {'料號':[],'barcode':[],'y料號':[],'品名':[],'神腦經銷價':[],
         't1未稅':[],'t1含稅':[],'t2七階':[],'市場售價':[],'t2未稅':[],'供應商':[],'日期':[]}
    for col, val in hd.items():
        v = safe_str(val)
        vf = v.split('\n')[0]  # first line only
        if vf in ('料號','A料號') or ('料號' in vf and 'Y料號' not in vf and '代號' not in vf):
            m['料號'].append(col)
        if '國際條碼' in v or 'Barcode' in v or 'barcode' in v:
            m['barcode'].append(col)
        if 'Y料號' in v or 'Ｙ料號' in v:
            m['y料號'].append(col)
        if '品名' in v:
            m['品名'].append(col)
        if '神腦經銷價' in v or '神腦' in vf:
            m['神腦經銷價'].append(col)
        if 'T1成本' in v or ('未稅' in v and 'T1' in v) or vf == '未稅':
            m['t1未稅'].append(col)
        if '進貨成本' in v or ('T1價格' in v and '含稅' in v):
            m['t1含稅'].append(col)
        if ('七階' in v or ('T2' in v and '寄售' in v and '含稅' in v)) and '優鋭' not in v:
            m['t2七階'].append(col)
        if '市場售價' in v or '優鋭七階' in v or 'MSRP' in v:
            m['市場售價'].append(col)
        if 'T2' in v and '優鋭' in v and '未稅' in v:
            m['t2未稅'].append(col)
        if '供應商' in v:
            m['供應商'].append(col)
        if '更新日期' in v:
            m['日期'].append(col)
    return m

def wk_make_row(cells, cm, date1904=True):
    品名 = get_val(cells, cm['品名'])
    if not 品名: return None
    s = safe_str(品名)
    if not s or s.startswith('T1') or s.startswith('T2'): return None

    日期 = get_val(cells, cm['日期'])
    if isinstance(日期, datetime.datetime): 日期 = 日期.strftime('%Y-%m-%d')
    elif isinstance(日期, (int,float)) and date1904: 日期 = excel_date_1904(日期)
    elif isinstance(日期, (int,float)): 日期 = excel_date_1900(日期)
    elif 日期: 日期 = safe_str(日期)

    料號    = get_val(cells, cm['料號'])
    bc      = get_val(cells, cm['barcode'])
    y料號   = get_val(cells, cm['y料號'])
    神腦    = get_val(cells, cm['神腦經銷價'])
    t1未稅  = get_val(cells, cm['t1未稅'])
    t1含稅  = get_val(cells, cm['t1含稅'])
    t2七階  = get_val(cells, cm['t2七階'])
    市場    = get_val(cells, cm['市場售價'])
    t2未稅  = get_val(cells, cm['t2未稅'])
    供應商  = get_val(cells, cm['供應商'])

    r料號 = safe_str(料號)
    if not r料號 or r料號 == '#N/A': r料號 = ''

    return {
        '供應商':    safe_str(供應商),
        '料號':     r料號,
        'barcode':  safe_str(bc).replace('.0',''),
        'y料號':    safe_str(y料號).replace('.0',''),
        '品名':     s,
        '神腦經銷價': safe_int(神腦),
        't1未稅':   safe_int(t1未稅),
        't1含稅':   safe_int(t1含稅),
        't2七階':   safe_int(t2七階),
        '市場售價':  safe_int(市場),
        't2未稅':   safe_int(t2未稅),
        '更新日期':  日期 or '',
    }

def parse_weiku_openpyxl(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except:
        return None
    result = {}
    for sname in wb.sheetnames:
        if sname in WEIKU_SKIP: continue
        ws = wb[sname]
        if ws.sheet_state == 'hidden':
            print(f"    ⚠️  分頁「{sname}」是隱藏分頁，已略過（若這是正式分類請取消隱藏）")
            continue
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True, max_col=25))
        cm, dstart = None, 0
        for i, row in enumerate(rows[:3]):
            hd = {chr(65+j): v for j,v in enumerate(row) if v is not None}
            if wk_is_header(hd): cm = wk_detect_cols(hd); dstart = i+1; break
        if not cm:
            if len(rows) > 0:
                print(f"    ⚠️  分頁「{sname}」有 {len(rows)} 列資料但找不到表頭，已略過（請確認欄位標題是否正常）")
            continue
        prods = []
        for row in rows[dstart:]:
            cells = {}
            for j,v in enumerate(row):
                if isinstance(v, datetime.datetime): v = v.strftime('%Y-%m-%d')
                cells[chr(65+j)] = v
            r = wk_make_row(cells, cm, date1904=False)
            if r: prods.append(r)
        if prods: result[WEIKU_RENAME.get(sname, sname)] = prods
    return result

def parse_weiku_xml(path):
    result = {}
    with zipfile.ZipFile(path, 'r') as z:
        shared = parse_shared_strings(z)
        sheet_map = get_sheet_map(z)
        for name, (sf, state) in sheet_map.items():
            if name in WEIKU_SKIP: continue
            if state == 'hidden':
                print(f"    ⚠️  分頁「{name}」是隱藏分頁，已略過（若這是正式分類請取消隱藏）")
                continue
            rows_raw = parse_sheet_xml(z, sf, shared)
            cm, dstart = None, 0
            for i, cells in enumerate(rows_raw[:3]):
                if wk_is_header(cells): cm = wk_detect_cols(cells); dstart = i+1; break
            if not cm:
                if len(rows_raw) > 0:
                    print(f"    ⚠️  分頁「{name}」有 {len(rows_raw)} 列資料但找不到表頭，已略過（請確認欄位標題是否正常）")
                continue
            prods = []
            for cells in rows_raw[dstart:]:
                r = wk_make_row(cells, cm, date1904=True)
                if r: prods.append(r)
            if prods: result[WEIKU_RENAME.get(name, name)] = prods
    return result

# ─── Process files ────────────────────────────────────────────────────────────
def process_folder(folder, parse_openpyxl_fn, parse_xml_fn, skip_set):
    all_data = {}
    xlsx_files = sorted([f for f in folder.glob('*.xlsx') if not f.name.startswith('~$')])
    if not xlsx_files:
        print(f"  ⚠️ {folder.name}/ 沒有 xlsx 檔案")
        return all_data
    for path in xlsx_files:
        print(f"\n  處理：{path.name}")
        result = parse_openpyxl_fn(path)
        if result is None:
            print(f"    → openpyxl 失敗，改用 XML 解析")
            result = parse_xml_fn(path)
        if not result:
            print(f"    ⚠️ 無有效資料"); continue
        for cat, prods in result.items():
            key = cat
            if key in all_data: key = cat + f'（{path.stem[:10]}）'
            all_data[key] = prods
            print(f"    ✓ {cat}：{len(prods)} 筆")
    return all_data

print("=" * 50)
print("  配件產品目錄產生器")
print("=" * 50)

print("\n【優仕配件】")
ys_data = process_folder(YOUSHIH_DIR, parse_youshih_openpyxl, parse_youshih_xml, YOUSHIH_SKIP) if YOUSHIH_DIR.exists() else {}

print("\n【暐固配件】")
wk_data = process_folder(WEIKU_DIR, parse_weiku_openpyxl, parse_weiku_xml, WEIKU_SKIP) if WEIKU_DIR.exists() else {}

ys_total = sum(len(v) for v in ys_data.values())
wk_total = sum(len(v) for v in wk_data.values())
print(f"\n優仕：{len(ys_data)} 分類，{ys_total:,} 筆")
print(f"暐固：{len(wk_data)} 分類，{wk_total:,} 筆")

# ─── Build HTML ───────────────────────────────────────────────────────────────
ys_js = json.dumps(ys_data, ensure_ascii=False)
wk_js = json.dumps(wk_data, ensure_ascii=False)
today = datetime.date.today().strftime('%Y-%m-%d')

html = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>配件產品目錄</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{--p:#1a56db;--pg:#059669;--bg:#f8fafc;--sf:#fff;--bd:#e2e8f0;--tx:#1e293b;--mu:#64748b;--hv:#f1f5f9}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Noto Sans TC",sans-serif;background:var(--bg);color:var(--tx);min-height:100vh}}

/* Header */
header{{background:linear-gradient(135deg,#1e40af,#1e3a8a);color:#fff;padding:18px 24px 14px;box-shadow:0 2px 8px rgba(0,0,0,.18)}}
header h1{{font-size:1.4rem;font-weight:700}}
header p{{font-size:.82rem;opacity:.75;margin-top:3px}}

/* Brand switcher */
.brand-bar{{background:#1e3a8a;display:flex;padding:0 24px;gap:4px}}
.brand-btn{{padding:10px 22px;font-size:.95rem;font-weight:600;cursor:pointer;border:none;background:transparent;color:rgba(255,255,255,.65);border-bottom:3px solid transparent;margin-bottom:0;transition:all .15s;letter-spacing:.01em}}
.brand-btn:hover{{color:#fff;background:rgba(255,255,255,.08)}}
.brand-btn.active{{color:#fff;border-bottom-color:#60a5fa}}
.brand-btn .bc{{font-size:.72rem;margin-left:5px;opacity:.8;font-weight:400}}

/* Toolbar */
.toolbar{{background:var(--sf);border-bottom:1px solid var(--bd);padding:11px 24px;display:flex;align-items:center;gap:10px;flex-wrap:wrap;position:sticky;top:0;z-index:100;box-shadow:0 1px 3px rgba(0,0,0,.06)}}
.sw{{position:relative;flex:1;min-width:240px;max-width:520px}}
.sw svg{{position:absolute;left:10px;top:50%;transform:translateY(-50%);color:var(--mu);pointer-events:none}}
#search{{width:100%;padding:8px 34px;border:1.5px solid var(--bd);border-radius:8px;font-size:.88rem;outline:none;transition:border-color .15s;background:var(--bg)}}
#search:focus{{border-color:var(--p);background:#fff}}
#clr{{position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;cursor:pointer;color:var(--mu);font-size:1rem;display:none}}
.scope{{display:flex;gap:5px;align-items:center;flex-wrap:wrap;font-size:.8rem;color:var(--mu)}}
.scope label{{display:flex;align-items:center;gap:3px;cursor:pointer;user-select:none;white-space:nowrap}}
.scope input{{accent-color:var(--p);width:13px;height:13px}}
.ib{{margin-left:auto;font-size:.8rem;color:var(--mu);white-space:nowrap}}

/* Search banner */
#banner{{display:none;background:#eff6ff;border-bottom:1px solid #bfdbfe;padding:7px 24px;font-size:.83rem;color:#1d4ed8;align-items:center;gap:8px}}
#banner.on{{display:flex}}
#banner button{{margin-left:auto;background:#1d4ed8;color:#fff;border:none;border-radius:5px;padding:3px 10px;cursor:pointer;font-size:.78rem}}

/* Category tabs */
.tabs-wrap{{background:var(--sf);border-bottom:2px solid var(--bd);padding:0 24px;overflow-x:auto;scrollbar-width:none}}
.tabs-wrap::-webkit-scrollbar{{display:none}}
.tabs{{display:flex;gap:1px;min-width:max-content}}
.tab{{padding:10px 14px;font-size:.85rem;font-weight:500;cursor:pointer;border:none;background:none;color:var(--mu);border-bottom:2px solid transparent;margin-bottom:-2px;transition:all .15s;white-space:nowrap;display:flex;align-items:center;gap:5px}}
.tab:hover{{color:var(--p);background:var(--hv);border-radius:5px 5px 0 0}}
.tab.active{{color:var(--p);border-bottom-color:var(--p);font-weight:600}}
.tab .ct{{background:#e0e7ff;color:#3730a3;font-size:.68rem;padding:1px 5px;border-radius:9999px;font-weight:600}}
.tab.active .ct{{background:var(--p);color:#fff}}

/* Table */
.content{{padding:14px 24px 40px}}
.tw{{background:var(--sf);border:1px solid var(--bd);border-radius:10px;overflow-x:auto;-webkit-overflow-scrolling:touch;box-shadow:0 1px 4px rgba(0,0,0,.06)}}
table{{width:100%;border-collapse:collapse;font-size:.83rem;min-width:700px}}
thead{{background:#f1f5f9}}
th{{padding:9px 12px;text-align:left;font-weight:600;font-size:.75rem;color:var(--mu);text-transform:uppercase;letter-spacing:.04em;border-bottom:1px solid var(--bd);cursor:pointer;user-select:none;white-space:nowrap}}
th:hover{{color:var(--p)}}
th .si{{margin-left:3px;opacity:.3}}
th.sorted .si{{opacity:1;color:var(--p)}}
td{{padding:8px 12px;border-bottom:1px solid #f1f5f9;vertical-align:middle}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:var(--hv)}}
.pno,.bc,.yno{{font-family:monospace;font-size:.77rem}}
.pno{{color:#7c3aed}}.bc{{color:var(--mu)}}.yno{{color:#0369a1}}
.nm{{font-weight:500;line-height:1.4;max-width:320px}}
.pr{{font-weight:600;color:#0f766e;white-space:nowrap}}
.ms{{color:var(--mu);white-space:nowrap}}
.vd{{display:inline-block;background:#fef3c7;color:#92400e;padding:1px 7px;border-radius:4px;font-size:.73rem;font-weight:500;white-space:nowrap}}
.dt{{color:var(--mu);font-size:.75rem;white-space:nowrap}}
.cat-b{{display:inline-block;background:#e0e7ff;color:#3730a3;padding:1px 7px;border-radius:4px;font-size:.73rem;font-weight:500;white-space:nowrap}}
mark{{background:#fef08a;border-radius:2px;padding:0 1px}}
.empty{{text-align:center;padding:50px 20px;color:var(--mu);font-size:.88rem}}

/* Pagination */
.pg{{display:flex;align-items:center;justify-content:center;gap:4px;padding:14px 0 0;flex-wrap:wrap}}
.pg button{{padding:5px 10px;border:1px solid var(--bd);background:#fff;border-radius:6px;cursor:pointer;font-size:.8rem;color:var(--tx);transition:all .15s}}
.pg button:hover{{border-color:var(--p);color:var(--p)}}
.pg button.on{{background:var(--p);color:#fff;border-color:var(--p)}}
.pg button:disabled{{opacity:.3;cursor:default}}
.pg .pi{{font-size:.78rem;color:var(--mu);padding:0 6px}}

/* Mobile */
@media(max-width:600px){{
  header{{padding:14px 16px}}
  header h1{{font-size:1.3rem}}
  header p{{font-size:.75rem}}
  .brand-bar{{padding:0 16px}}
  .brand-btn{{padding:10px 18px;font-size:.85rem}}
  .toolbar{{padding:10px 16px;gap:8px}}
  .content{{padding:10px 16px 40px}}
  .scope{{font-size:.78rem;gap:6px}}
  table{{font-size:.75rem}}
  th{{padding:7px 8px;font-size:.68rem}}
  td{{padding:6px 8px}}
  .nm{{max-width:160px}}
}}

/* ── Compare modal ─────────────────────────────────────────────────────────── */
.modal-bg{{position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:200;align-items:center;justify-content:center;display:none}}
.modal-bg.open{{display:flex}}
.modal{{background:#fff;border-radius:12px;padding:24px 28px;width:min(580px,95vw);max-height:88vh;overflow-y:auto;box-shadow:0 8px 32px rgba(0,0,0,.2)}}
.modal h2{{font-size:1rem;font-weight:700;margin-bottom:16px;color:var(--tx);display:flex;align-items:center;justify-content:space-between}}
.modal h2 button{{background:none;border:none;cursor:pointer;font-size:.85rem;color:var(--mu);padding:2px 6px}}
.drop-area{{border:2px dashed var(--bd);border-radius:8px;padding:26px;text-align:center;cursor:pointer;color:var(--mu);font-size:.85rem;transition:border-color .15s;margin-bottom:10px}}
.drop-area:hover,.drop-area.over{{border-color:var(--p);color:var(--p)}}
.drop-area input{{display:none}}
.or-sep{{text-align:center;color:var(--mu);font-size:.76rem;margin:8px 0;position:relative}}
.or-sep::before,.or-sep::after{{content:'';position:absolute;top:50%;width:43%;border-top:1px solid var(--bd)}}
.or-sep::before{{left:0}}.or-sep::after{{right:0}}
#paste-area{{width:100%;height:88px;resize:vertical;border:1.5px solid var(--bd);border-radius:7px;padding:8px;font-size:.82rem;font-family:monospace;outline:none;box-sizing:border-box}}
#paste-area:focus{{border-color:var(--p)}}
.modal-hint{{font-size:.75rem;color:var(--mu);margin:6px 0 14px;line-height:1.55}}
.modal-actions{{display:flex;gap:8px;flex-wrap:wrap}}
.btn-prim{{background:var(--p);color:#fff;border:none;border-radius:7px;padding:8px 18px;cursor:pointer;font-size:.84rem;font-weight:600}}
.btn-prim:hover{{background:#1740b0}}
.btn-sec2{{background:#fff;color:var(--tx);border:1.5px solid var(--bd);border-radius:7px;padding:8px 18px;cursor:pointer;font-size:.84rem;font-weight:500}}
.btn-sec2:hover:not(:disabled){{border-color:var(--p);color:var(--p)}}
.btn-sec2:disabled{{opacity:.35;cursor:default}}
.cmp-summary{{margin:14px 0 8px;font-size:.84rem;display:flex;gap:14px;flex-wrap:wrap;align-items:center}}
.cmp-hit{{color:#059669;font-weight:700}}.cmp-miss{{color:#dc2626;font-weight:700}}
.cmp-wrap{{max-height:260px;overflow-y:auto;border:1px solid var(--bd);border-radius:7px;margin-top:4px}}
.cmp-tbl{{width:100%;border-collapse:collapse;font-size:.78rem}}
.cmp-tbl th{{background:#f1f5f9;padding:6px 10px;text-align:left;font-size:.7rem;font-weight:600;color:var(--mu);text-transform:uppercase;letter-spacing:.03em;position:sticky;top:0;border-bottom:1px solid var(--bd)}}
.cmp-tbl td{{padding:6px 10px;border-top:1px solid #f1f5f9}}
.cmp-tbl tr:first-child td{{border-top:none}}
.cmp-tbl tr:hover td{{background:var(--hv)}}
.bg-hit{{background:#d1fae5;color:#065f46;padding:1px 7px;border-radius:4px;font-size:.7rem;font-weight:600;white-space:nowrap;display:inline-block}}
.bg-miss{{background:#fee2e2;color:#991b1b;padding:1px 7px;border-radius:4px;font-size:.7rem;font-weight:600;white-space:nowrap;display:inline-block}}
</style>
</head>
<body>
<header>
  <h1>配件產品目錄</h1>
  <p>優仕 {ys_total:,} 筆 · 暐固 {wk_total:,} 筆 · 最後更新：{today}</p>
</header>

<div class="brand-bar">
  <button class="brand-btn active" id="b-ys" onclick="setBrand('ys')">
    優仕配件 <span class="bc">{ys_total:,} 筆</span>
  </button>
  <button class="brand-btn" id="b-wk" onclick="setBrand('wk')">
    暐固配件 <span class="bc">{wk_total:,} 筆</span>
  </button>
</div>

<div class="toolbar">
  <div class="sw">
    <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.2"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>
    <input type="text" id="search" placeholder="搜尋品名、料號、國條/Barcode…（跨所有分類）">
    <button id="clr">✕</button>
  </div>
  <div class="scope">
    <span>欄位：</span>
    <label><input type="checkbox" id="sc-nm" checked> 品名</label>
    <label><input type="checkbox" id="sc-pno" checked> 料號</label>
    <label><input type="checkbox" id="sc-bc" checked> 國條/Barcode</label>
    <label><input type="checkbox" id="sc-vd" checked> 供應商/廠商</label>
    <label id="sc-yno-lbl" style="display:none"><input type="checkbox" id="sc-yno" checked> Y料號</label>
  </div>
  <span class="ib" id="ib"></span>
  <button class="btn-sec2" style="padding:7px 13px;font-size:.82rem" onclick="openCompare()">📥 比對清單</button>
</div>

<div id="banner">
  <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>
  <span>跨分類搜尋：<strong id="bq"></strong> — <strong id="bc2"></strong> 筆，來自 <strong id="bcats"></strong> 個分類</span>
  <button onclick="clrQ()">清除搜尋</button>
</div>

<div class="tabs-wrap"><div class="tabs" id="tabs"></div></div>
<div class="content">
  <div class="tw"><table><thead id="thead"><tr></tr></thead><tbody id="tbody"></tbody></table></div>
  <div class="pg" id="pg"></div>
</div>

<!-- ── Compare modal ─────────────────────────────────────────────────────── -->
<div class="modal-bg" id="cmp-modal" onclick="if(event.target===this)closeCompare()">
  <div class="modal">
    <h2>📥 比對清單
      <button onclick="closeCompare()">✕ 關閉</button>
    </h2>

    <div class="drop-area" id="drop-area"
      onclick="document.getElementById('cmp-file').click()"
      ondragover="event.preventDefault();this.classList.add('over')"
      ondragleave="this.classList.remove('over')"
      ondrop="event.preventDefault();this.classList.remove('over');handleFile(event.dataTransfer.files[0])">
      <input type="file" id="cmp-file" accept=".csv,.xlsx,.xls,.txt"
        onchange="handleFile(this.files[0])">
      <svg width="26" height="26" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5"
        style="margin:0 auto 8px;display:block;opacity:.4">
        <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>
        <polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/>
      </svg>
      <div id="drop-label">點擊上傳 CSV 或 Excel 檔，或拖曳至此</div>
    </div>

    <div class="or-sep">或手動貼上代碼</div>
    <textarea id="paste-area" placeholder="每行貼一個料號或國條/Barcode，或整批貼上（自動偵測欄位）"></textarea>
    <div class="modal-hint">
      支援格式：<strong>CSV / Excel（.xlsx/.xls）/ 純文字</strong>，自動偵測料號或國條所在欄位。<br>
      同時搜尋優仕＋暐固兩個品牌，結果可匯出為 CSV。
    </div>

    <div class="modal-actions">
      <button class="btn-prim" onclick="runCompare()">🔍 開始比對</button>
      <button class="btn-sec2" id="dl-csv-btn" onclick="downloadCompareCsv()" disabled>📥 匯出結果 CSV</button>
      <button class="btn-sec2" onclick="resetCompare()">清除</button>
    </div>
    <div id="cmp-result"></div>
  </div>
</div>

<script>
// ── Data ──────────────────────────────────────────────────────────────────────
const YS = {ys_js};
const WK = {wk_js};

// Flatten for cross-search
const YS_ALL=[],WK_ALL=[];
Object.entries(YS).forEach(([c,rs])=>rs.forEach(r=>YS_ALL.push({{...r,__c:c}})));
Object.entries(WK).forEach(([c,rs])=>rs.forEach(r=>WK_ALL.push({{...r,__c:c}})));

// ── State ─────────────────────────────────────────────────────────────────────
let brand='ys', cur='', q='', sc='', sd=1, pg=1;
const PS=100;

// ── Brand columns config ──────────────────────────────────────────────────────
const COLS = {{
  ys: [
    {{key:'廠商',    label:'廠商',           cls:'',    fmt:v=>`<span class="vd">${{esc(v||'—')}}</span>`,raw:true}},
    {{key:'料號',    label:'優仕料號',       cls:'pno', fmt:v=>v||'—'}},
    {{key:'barcode', label:'國條/Barcode',   cls:'bc',  fmt:v=>v||'—'}},
    {{key:'品名',    label:'品名',           cls:'nm',  fmt:v=>v}},
    {{key:'進價',    label:'進價（未稅）',    cls:'pr',  fmt:fmtP}},
    {{key:'售價',    label:'建議售價（含）',  cls:'ms',  fmt:fmtP}},
    {{key:'建檔日期',label:'建檔日期',        cls:'dt',  fmt:v=>v||'—'}},
  ],
  wk: [
    {{key:'供應商',   label:'供應商',                  cls:'',   fmt:v=>`<span class="vd">${{esc(v||'—')}}</span>`,raw:true}},
    {{key:'料號',    label:'暐固料號',                 cls:'pno',fmt:v=>v||'—'}},
    {{key:'barcode', label:'國際條碼/Barcode',         cls:'bc', fmt:v=>v||'—'}},
    {{key:'y料號',   label:'Y料號',                   cls:'yno',fmt:v=>v||'—'}},
    {{key:'品名',    label:'品名',                    cls:'nm', fmt:v=>v}},
    {{key:'t1未稅',  label:'T1成本（未稅）',           cls:'pr', fmt:fmtP}},
    {{key:'t1含稅',  label:'T1價格（含稅）暐固進貨成本',cls:'ms', fmt:fmtP}},
    {{key:'t2七階',  label:'T2寄售成本（含）暐固七階',  cls:'ms', fmt:fmtP}},
    {{key:'市場售價', label:'市場售價 優鋭七階',         cls:'ms', fmt:fmtP}},
    {{key:'t2未稅',  label:'T2出優鋭成本（未稅）',      cls:'ms', fmt:fmtP}},
    {{key:'更新日期', label:'更新日期',                 cls:'dt', fmt:v=>v||'—'}},
  ]
}};
const SEARCH_KEYS = {{
  ys: {{nm:'品名', pno:'料號', bc:'barcode', vd:'廠商'}},
  wk: {{nm:'品名', pno:'料號', bc:'barcode', vd:'供應商', yno:'y料號'}},
}};

// ── Helpers ───────────────────────────────────────────────────────────────────
function esc(s){{return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')}}
function hi(t,q){{
  if(!q)return esc(t);
  const r=new RegExp('('+q.replace(/[.*+?^${{}}()|[\\]\\\\]/g,'\\\\$&')+')','gi');
  return esc(t).replace(r,'<mark>$1</mark>');
}}
function fmtP(v){{const n=parseInt(v);return isNaN(n)?'—':'$'+n.toLocaleString()}}
function data(){{return brand==='ys'?YS:WK}}
function allRows(){{return brand==='ys'?YS_ALL:WK_ALL}}
function scope(){{return{{nm:document.getElementById('sc-nm').checked,pno:document.getElementById('sc-pno').checked,bc:document.getElementById('sc-bc').checked,vd:document.getElementById('sc-vd').checked,yno:(document.getElementById('sc-yno')||{{}}).checked}}}}
function matchRow(r,lq,s,sk){{
  return(s.nm&&(r[sk.nm]||'').toLowerCase().includes(lq))||
        (s.pno&&(r[sk.pno]||'').toLowerCase().includes(lq))||
        (s.bc&&(r[sk.bc]||'').toLowerCase().includes(lq))||
        (s.vd&&(r[sk.vd]||'').toLowerCase().includes(lq))||
        (s.yno&&sk.yno&&(r[sk.yno]||'').toLowerCase().includes(lq));
}}

// ── Brand switch ──────────────────────────────────────────────────────────────
function setBrand(b){{
  brand=b; pg=1;
  document.getElementById('b-ys').classList.toggle('active',b==='ys');
  document.getElementById('b-wk').classList.toggle('active',b==='wk');
  const ynoLbl=document.getElementById('sc-yno-lbl');
  if(ynoLbl)ynoLbl.style.display=b==='wk'?'':'none';
  const D=data();
  cur=Object.keys(D)[0]||'';
  buildTabs(); buildThead(); render();
}}

// ── Tabs ──────────────────────────────────────────────────────────────────────
function buildTabs(){{
  const D=data(), el=document.getElementById('tabs');
  el.innerHTML='';
  Object.entries(D).forEach(([c,rs])=>{{
    const b=document.createElement('button');
    b.className='tab'; b.dataset.cat=c;
    b.innerHTML=esc(c)+' <span class="ct">'+rs.length.toLocaleString()+'</span>';
    b.onclick=()=>{{cur=c;pg=1;render()}};
    el.appendChild(b);
  }});
}}

// ── Thead ─────────────────────────────────────────────────────────────────────
function buildThead(withCat){{
  const cols=COLS[brand];
  const el=document.getElementById('thead').querySelector('tr');
  const catTh=withCat?`<th data-col="__c" onclick="srt('__c')">產品分類<span class="si">↕</span></th>`:'';
  el.innerHTML=catTh+cols.map(c=>`<th data-col="${{c.key}}" onclick="srt('${{c.key}}')">${{c.label}}<span class="si">↕</span></th>`).join('');
}}

// ── Render ─────────────────────────────────────────────────────────────────────
function render(){{
  const D=data(), sk=SEARCH_KEYS[brand], cols=COLS[brand], s=scope();
  const lq=q.toLowerCase();
  let rows, catCol=false;

  if(q){{
    rows=allRows().filter(r=>matchRow(r,lq,s,sk));
    catCol=true;
  }} else {{
    rows=D[cur]||[];
  }}

  // Sort
  if(sc){{
    rows=[...rows].sort((a,b)=>{{
      let va=a[sc]||'',vb=b[sc]||'';
      const numCols=['進價','售價','t1未稅','t1含稅','t2七階','市場售價'];
      if(numCols.includes(sc))return((parseFloat(va)||0)-(parseFloat(vb)||0))*sd;
      return String(va).localeCompare(String(vb),'zh-TW')*sd;
    }});
  }}

  const tot=rows.length, tp=Math.max(1,Math.ceil(tot/PS));
  if(pg>tp)pg=tp;
  const sl=rows.slice((pg-1)*PS,pg*PS);

  // Update tabs
  document.querySelectorAll('.tab').forEach(t=>t.classList.toggle('active',!q&&t.dataset.cat===cur));

  // Banner
  const bn=document.getElementById('banner');
  if(q){{
    document.getElementById('bq').textContent=q;
    document.getElementById('bc2').textContent=tot.toLocaleString();
    document.getElementById('bcats').textContent=new Set(rows.map(r=>r.__c)).size;
    bn.classList.add('on');
  }} else bn.classList.remove('on');

  // Info
  document.getElementById('ib').textContent=q?tot.toLocaleString()+' 筆':'顯示 '+tot.toLocaleString()+' 筆';

  // Thead
  buildThead(catCol);
  // Re-apply sort indicator
  if(sc) document.querySelectorAll('th').forEach(t=>{{
    const m=t.dataset.col===sc;
    t.classList.toggle('sorted',m);
    const si=t.querySelector('.si');
    if(si)si.textContent=m?(sd===1?' ↑':' ↓'):'↕';
  }});

  // Tbody
  const tb=document.getElementById('tbody');
  if(!sl.length){{
    tb.innerHTML=`<tr><td colspan="${{(catCol?1:0)+cols.length}}" class="empty">查無符合「${{esc(q)}}」的產品</td></tr>`;
  }} else {{
    tb.innerHTML=sl.map(r=>{{
      const catCell=catCol?`<td><span class="cat-b">${{esc(r.__c)}}</span></td>`:'';
      const cells=cols.map(c=>{{
        const v=r[c.key];
        let rendered;
        if(c.raw){{rendered=c.fmt(v||'')}}
        else{{rendered=`<span class="${{c.cls}}">${{hi(c.fmt(v||''),q)}}</span>`}}
        return`<td>${{rendered}}</td>`;
      }}).join('');
      return`<tr>${{catCell}}${{cells}}</tr>`;
    }}).join('');
  }}

  // Pagination
  const el=document.getElementById('pg');
  if(tp<=1){{el.innerHTML='';return}}
  let b=`<button onclick="gp(${{pg-1}})" ${{pg===1?'disabled':''}}>‹ 上頁</button>`;
  const rng=[];
  for(let i=1;i<=tp;i++){{if(i===1||i===tp||(i>=pg-2&&i<=pg+2))rng.push(i);else if(rng[rng.length-1]!=='…')rng.push('…')}}
  rng.forEach(r=>{{if(r==='…')b+=`<button disabled>…</button>`;else b+=`<button class="${{r===pg?'on':''}}" onclick="gp(${{r}})">${{r}}</button>`}});
  b+=`<button onclick="gp(${{pg+1}})" ${{pg===tp?'disabled':''}}>下頁 ›</button>`;
  b+=`<span class="pi">第 ${{pg}} / ${{tp}} 頁</span>`;
  el.innerHTML=b;
}}

function gp(p){{pg=p;render();window.scrollTo(0,0)}}

function srt(col){{
  sd=(sc===col)?sd*-1:1; sc=col; pg=1; render();
}}

function clrQ(){{
  document.getElementById('search').value=''; q=''; pg=1;
  document.getElementById('clr').style.display='none'; render();
}}

// Search
let tmr;
document.getElementById('search').addEventListener('input',e=>{{
  clearTimeout(tmr);
  tmr=setTimeout(()=>{{q=e.target.value.trim();pg=1;document.getElementById('clr').style.display=q?'block':'none';render()}},180);
}});
document.getElementById('clr').addEventListener('click',clrQ);
['sc-nm','sc-pno','sc-bc','sc-vd','sc-yno'].forEach(id=>{{const el=document.getElementById(id);if(el)el.addEventListener('change',()=>{{if(q){{pg=1;render()}}}});}});

// ── Compare feature ───────────────────────────────────────────────────────────
let _cmpData = null;   // last compare results
let _parsedCodes = []; // codes parsed from file/paste

// Build fast lookup: code (lower) -> {{品牌, 分類, 品名, 料號, barcode, y料號}}
function buildLookup() {{
  const lk = {{}};
  function add(r, 品牌, 分類) {{
    const entry = {{品牌, 分類, 品名:r['品名']||'', 料號:r['料號']||'', barcode:r['barcode']||'', 'y料號':r['y料號']||''}};
    ['料號','barcode','y料號'].forEach(k=>{{ if(r[k]) lk[String(r[k]).trim().toLowerCase()] = entry; }});
  }}
  YS_ALL.forEach(r=>add(r,'優仕',r.__c));
  WK_ALL.forEach(r=>add(r,'暐固',r.__c));
  return lk;
}}

// Parse a 2-D array (rows of cells) → list of candidate codes
// Detects which column(s) look like 料號 / barcode
function extractCodes(rows) {{
  if(!rows.length) return [];
  const header = rows[0].map(v=>String(v||'').trim().toLowerCase());
  // Preferred columns by header keyword match
  const pnoIdx = header.findIndex(h=>h.includes('料號')||h.includes('pno')||h.includes('sku'));
  const bcIdx  = header.findIndex(h=>h.includes('barcode')||h.includes('國條')||h.includes('條碼'));
  const codes = [];
  const dataRows = (pnoIdx>=0||bcIdx>=0) ? rows.slice(1) : rows; // skip header if found
  dataRows.forEach(row=>{{
    if(pnoIdx>=0 && row[pnoIdx]) codes.push(String(row[pnoIdx]).trim());
    if(bcIdx >=0 && row[bcIdx])  codes.push(String(row[bcIdx]).trim());
    // If no header found, treat every non-empty cell in col 0 as a code
    if(pnoIdx<0 && bcIdx<0 && row[0]) codes.push(String(row[0]).trim());
  }});
  return [...new Set(codes.filter(Boolean))];
}}

// Handle file upload (CSV or Excel)
function handleFile(file) {{
  if(!file) return;
  document.getElementById('drop-label').textContent = '✓ 已載入：' + file.name;
  const reader = new FileReader();
  const ext = file.name.split('.').pop().toLowerCase();
  if(ext==='xlsx'||ext==='xls') {{
    reader.onload = e=>{{
      const wb = XLSX.read(new Uint8Array(e.target.result), {{type:'array'}});
      const ws = wb.Sheets[wb.SheetNames[0]];
      const rows = XLSX.utils.sheet_to_json(ws, {{header:1, defval:''}});
      _parsedCodes = extractCodes(rows);
      document.getElementById('paste-area').value = _parsedCodes.join('\n');
    }};
    reader.readAsArrayBuffer(file);
  }} else {{
    reader.onload = e=>{{
      const wb = XLSX.read(e.target.result, {{type:'string'}});
      const ws = wb.Sheets[wb.SheetNames[0]];
      const rows = XLSX.utils.sheet_to_json(ws, {{header:1, defval:''}});
      _parsedCodes = extractCodes(rows);
      document.getElementById('paste-area').value = _parsedCodes.join('\n');
    }};
    reader.readAsText(file, 'UTF-8');
  }}
}}

function openCompare() {{
  document.getElementById('cmp-modal').classList.add('open');
}}
function closeCompare() {{
  document.getElementById('cmp-modal').classList.remove('open');
}}
function resetCompare() {{
  _cmpData = null; _parsedCodes = [];
  document.getElementById('paste-area').value = '';
  document.getElementById('drop-label').textContent = '點擊上傳 CSV 或 Excel 檔，或拖曳至此';
  document.getElementById('cmp-file').value = '';
  document.getElementById('cmp-result').innerHTML = '';
  document.getElementById('dl-csv-btn').disabled = true;
}}

function runCompare() {{
  const raw = document.getElementById('paste-area').value.trim();
  const codes = raw ? [...new Set(raw.split(/[\n,\t]+/).map(s=>s.trim()).filter(Boolean))] : _parsedCodes;
  if(!codes.length) {{ alert('請先上傳檔案或貼上代碼'); return; }}
  const lk = buildLookup();
  _cmpData = codes.map(c=>{{
    const m = lk[c.toLowerCase()];
    return m ? {{查詢代碼:c, 狀態:'已收錄', 品牌:m.品牌, 分類:m.分類, 品名:m.品名, 料號:m.料號, barcode:m.barcode, 'y料號':m['y料號']}}
             : {{查詢代碼:c, 狀態:'未收錄', 品牌:'', 分類:'', 品名:'', 料號:'', barcode:'', 'y料號':''}};
  }});
  const hit = _cmpData.filter(r=>r.狀態==='已收錄').length;
  const miss = _cmpData.length - hit;
  document.getElementById('dl-csv-btn').disabled = false;
  const rows = _cmpData.map(r=>`
    <tr>
      <td>${{esc(r.查詢代碼)}}</td>
      <td><span class="${{r.狀態==='已收錄'?'bg-hit':'bg-miss'}}">${{r.狀態}}</span></td>
      <td>${{esc(r.品牌)}}</td>
      <td><span class="cat-b" style="display:${{r.分類?'inline-block':'none'}}">${{esc(r.分類)}}</span></td>
      <td style="max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${{esc(r.品名)}}">${{esc(r.品名)}}</td>
      <td class="pno">${{esc(r.料號)}}</td>
    </tr>`).join('');
  document.getElementById('cmp-result').innerHTML = `
    <div class="cmp-summary">
      共 <strong>${{_cmpData.length}}</strong> 筆 ·
      <span class="cmp-hit">✓ 已收錄 ${{hit}} 筆</span>
      <span class="cmp-miss">✗ 未收錄 ${{miss}} 筆</span>
    </div>
    <div class="cmp-wrap">
      <table class="cmp-tbl">
        <thead><tr><th>查詢代碼</th><th>狀態</th><th>品牌</th><th>分類</th><th>品名</th><th>料號</th></tr></thead>
        <tbody>${{rows}}</tbody>
      </table>
    </div>`;
}}

function downloadCompareCsv() {{
  if(!_cmpData) return;
  const cols = ['查詢代碼','狀態','品牌','分類','品名','料號','barcode','y料號'];
  const lines = [cols, ..._cmpData.map(r=>cols.map(k=>`"${{String(r[k]||'').replace(/"/g,'""')}}"`))]
    .map(r=>r.join(',')).join('\r\n');
  const blob = new Blob(['﻿'+lines], {{type:'text/csv;charset=utf-8'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  const d = new Date().toISOString().slice(0,10);
  a.download = `比對結果_${{d}}.csv`;
  a.click();
}}

// Init
setBrand('ys');
</script>
</body>
</html>"""

with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\n✅ 完成！已產生：{OUTPUT_HTML}")
print(f"   ({len(html.encode('utf-8'))//1024} KB)")
